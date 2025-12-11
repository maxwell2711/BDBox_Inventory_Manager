import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime


class InventoryManager:
    """Manages inventory data extraction, storage, and analysis across multiple Excel sheets."""
    
    def __init__(self, output_file_path: str = "Inventory_Analysis.xlsx"):
        """Initialize the Inventory Manager.
        
        Args:
            output_file_path: Path to the Excel file for storing analysis results
        """
        self.output_file = output_file_path
        self.sheet_names = {
            'history': 'Inventory History',
            'differences': 'Sales Differences',
            'average': 'Average Use',
            'predictions': 'Current Inventory & Predictions'
        }
    
    def extract_columns(self, file_path: str, column_headers: list) -> dict:
        """Extract specific columns from an Excel file based on header titles.
        
        Args:
            file_path: Path to the input Excel file
            column_headers: List of column header names to extract
            
        Returns:
            Dictionary with header names as keys and lists of values as values
            
        Raises:
            ValueError: If any column header is not found in the Excel file
        """
        df = pd.read_excel(file_path)
        
        # Verify all requested columns exist
        missing_columns = [col for col in column_headers if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Columns not found in Excel file: {missing_columns}")
        
        # Extract requested columns
        extracted_data = {col: df[col].tolist() for col in column_headers}
        return extracted_data
    
    def _get_or_create_sheet(self, wb, sheet_name: str):
        """Get an existing sheet or create a new one.
        
        Args:
            wb: Workbook object
            sheet_name: Name of the sheet
            
        Returns:
            Worksheet object
        """
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        else:
            return wb.create_sheet(sheet_name)
    
    def _load_or_create_workbook(self):
        """Load existing workbook or create a new one.
        
        Returns:
            Workbook object
        """
        output_path = Path(self.output_file)
        if output_path.exists():
            return load_workbook(output_path)
        else:
            wb = Workbook()
            # Remove the default sheet if it exists
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            return wb
    
    def _organize_sheets(self, wb):
        """Organize sheets in the desired order and remove any default sheets.
        
        Order: Current Inventory & Predictions, Inventory History, Sales Differences, Average Use
        
        Args:
            wb: Workbook object
        """
        # Define desired sheet order
        desired_order = [
            self.sheet_names['predictions'],
            self.sheet_names['history'],
            self.sheet_names['differences'],
            self.sheet_names['average']
        ]
        
        # Remove any default or unwanted sheets
        for sheet_name in wb.sheetnames[:]:
            if sheet_name not in desired_order:
                wb.remove(wb[sheet_name])
        
        # Reorder sheets according to desired order
        for idx, sheet_name in enumerate(desired_order):
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                wb.move_sheet(sheet, offset=idx - wb.index(sheet))
    
    def _get_existing_labels(self, ws):
        """Extract existing labels from column A of a worksheet.
        
        Args:
            ws: Worksheet object
            
        Returns:
            List of label values (excluding 'Label' header)
        """
        return [cell.value for cell in ws['A'][1:] if cell.value and cell.value != 'Label']
    
    def _merge_labels(self, existing_labels: list, new_labels: list) -> list:
        """Merge existing and new labels while maintaining order and removing duplicates.
        
        Args:
            existing_labels: List of existing labels
            new_labels: List of new labels to add
            
        Returns:
            Merged list without duplicates
        """
        return list(dict.fromkeys(existing_labels + new_labels))
    
    def _write_labels_to_column(self, ws, labels: list, column: str = 'A'):
        """Write labels to a column in the worksheet with 'Label' header.
        
        Args:
            ws: Worksheet object
            labels: List of labels to write
            column: Column letter (default 'A')
        """
        ws[f'{column}1'] = 'Label'
        for idx, label in enumerate(labels, start=2):
            ws[f'{column}{idx}'] = label
    
    def _add_inventory_column(self, ws, all_labels: list, label_to_stock: dict, column_header: str):
        """Add a new inventory column to the worksheet.
        
        Args:
            ws: Worksheet object
            all_labels: Complete merged list of all box labels in correct order
            label_to_stock: Dictionary mapping labels to their stock values
            column_header: Header for the new column (e.g., 'Sale 155' or 'Restock')
        """
        # Find the next available column
        next_col = ws.max_column + 1 if ws.max_column > 0 else 2
        col_letter = get_column_letter(next_col)
        
        # Write the column header
        ws[f'{col_letter}1'] = column_header
        
        # Write stock values, leaving blanks for new boxes
        for idx, label in enumerate(all_labels, start=2):
            if label in label_to_stock:
                ws[f'{col_letter}{idx}'] = label_to_stock[label]
    
    def _realign_existing_columns(self, ws, existing_labels: list, all_labels: list):
        """Realign all existing inventory columns when new labels are added.
        
        This ensures that stock values in previous columns stay aligned with their correct labels
        after new labels are inserted into the merged list.
        
        Args:
            ws: Worksheet object
            existing_labels: Original list of labels before merge
            all_labels: Merged list of labels after adding new ones
        """
        # Create a mapping of old positions to new positions
        label_to_new_row = {label: idx + 2 for idx, label in enumerate(all_labels)}
        
        # Get all inventory columns (columns B onwards)
        for col in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col)
            
            # Store all values first to avoid overwriting
            values_to_move = {}
            for idx, label in enumerate(existing_labels, start=2):
                cell_value = ws[f'{col_letter}{idx}'].value
                if cell_value is not None:
                    values_to_move[label] = cell_value
            
            # Clear the entire column except header
            for row in range(2, ws.max_row + 1):
                ws[f'{col_letter}{row}'].value = None
            
            # Write values back to new positions
            for label, value in values_to_move.items():
                new_row = label_to_new_row[label]
                ws[f'{col_letter}{new_row}'] = value
    
    def _update_inventory_history_internal(self, labels: list, stock_values: list, column_header: str):
        """Internal method to update inventory history with a new column.
        
        Args:
            labels: List of box labels from the input file
            stock_values: List of corresponding stock values
            column_header: Column header (e.g., 'Sale 155' or 'Restock')
        """
        output_path = Path(self.output_file)
        
        # Load or create the workbook
        wb = self._load_or_create_workbook()
        ws = self._get_or_create_sheet(wb, self.sheet_names['history'])
        
        # Get existing labels and merge with new ones
        existing_labels = self._get_existing_labels(ws) if ws.max_row > 0 else []
        all_labels = self._merge_labels(existing_labels, labels)
        
        # Realign existing columns if new labels were added
        if len(all_labels) > len(existing_labels):
            self._realign_existing_columns(ws, existing_labels, all_labels)
        
        # Write labels to column A
        self._write_labels_to_column(ws, all_labels)
        
        # Create a mapping of labels to stock values
        label_to_stock = dict(zip(labels, stock_values))
        
        # Add the new inventory column
        self._add_inventory_column(ws, all_labels, label_to_stock, column_header)
        
        # Organize sheets and remove defaults
        self._organize_sheets(wb)
        wb.save(output_path)
    
    def update_inventory_history(self, labels: list, stock_values: list, sale_number: str):
        """Update the Inventory History sheet with new sale data.
        
        Maintains alignment of box labels and backfills missing data for new boxes.
        
        Args:
            labels: List of box labels from the input file
            stock_values: List of corresponding stock values
            sale_number: Sale number to use as column header
        """
        self._update_inventory_history_internal(labels, stock_values, f'Sale {sale_number}')
    
    def _update_inventory_history_restock(self, labels: list, stock_values: list):
        """Update the Inventory History sheet with a restock entry.
        
        Args:
            labels: List of box labels from the input file
            stock_values: List of corresponding stock values
        """
        self._update_inventory_history_internal(labels, stock_values, 'Restock')
    
    def update_sales_differences(self):
        """Calculate and update the Sales Differences sheet.
        
        Computes differences between consecutive sales (or restock to sale if consecutive).
        Skips differences if sales are not consecutive.
        """
        output_path = Path(self.output_file)
        if not output_path.exists():
            raise FileNotFoundError(f"Output file {self.output_file} does not exist")
        
        wb = load_workbook(output_path)
        history_ws = wb[self.sheet_names['history']]
        
        # Get labels and sales data from history sheet (skip header row)
        labels = [cell.value for cell in history_ws['A'][1:] if cell.value and cell.value != 'Label']
        
        # Get all inventory columns (columns B onwards) with their metadata
        inventory_columns = []
        for col in range(2, history_ws.max_column + 1):
            col_letter = get_column_letter(col)
            header = history_ws[f'{col_letter}1'].value
            
            # Parse header to determine if it's a Sale or Restock
            if header.startswith('Sale '):
                sale_num = int(header.replace('Sale ', ''))
                inventory_columns.append((col, col_letter, header, 'sale', sale_num))
            elif header.startswith('Restock'):
                inventory_columns.append((col, col_letter, header, 'restock', None))
        
        # Create or update the differences sheet
        if self.sheet_names['differences'] in wb.sheetnames:
            diff_ws = wb[self.sheet_names['differences']]
            diff_ws.delete_rows(1, diff_ws.max_row)
        else:
            diff_ws = wb.create_sheet(self.sheet_names['differences'])
        
        # Write header and labels
        diff_ws['A1'] = 'Label'
        for idx, label in enumerate(labels, start=2):
            diff_ws[f'A{idx}'] = label
        
        # Calculate differences between consecutive entries
        diff_col_counter = 2  # Start from column B
        for i in range(len(inventory_columns) - 1):
            col1, col1_letter, header1, type1, sale_num1 = inventory_columns[i]
            col2, col2_letter, header2, type2, sale_num2 = inventory_columns[i + 1]
            
            # Determine if we should calculate difference
            should_calc_diff = False
            diff_header = None
            
            # If both are sales and consecutive, calculate difference
            if type1 == 'sale' and type2 == 'sale' and sale_num2 - sale_num1 == 1:
                should_calc_diff = True
                diff_header = f'Difference Sale {sale_num1} - Sale {sale_num2}'
            # If first is restock and second is sale, calculate difference
            elif type1 == 'restock' and type2 == 'sale':
                should_calc_diff = True
                diff_header = f'Difference Restock - Sale {sale_num2}'
            
            if not should_calc_diff:
                continue
            
            diff_col_letter = get_column_letter(diff_col_counter)
            diff_ws[f'{diff_col_letter}1'] = diff_header
            
            # Calculate differences
            for idx, label in enumerate(labels, start=2):
                val1 = history_ws[f'{col1_letter}{idx}'].value
                val2 = history_ws[f'{col2_letter}{idx}'].value
                
                # Only calculate if both values exist
                if val1 is not None and val2 is not None:
                    try:
                        diff = float(val1) - float(val2)
                        diff_ws[f'{diff_col_letter}{idx}'] = diff
                    except (ValueError, TypeError):
                        pass
            
            diff_col_counter += 1
        
        # Organize sheets and remove defaults
        self._organize_sheets(wb)
        wb.save(output_path)
    
    def update_average_use(self):
        """Calculate and update the Average Use sheet.
        
        Averages all difference columns from the Sales Differences sheet.
        """
        output_path = Path(self.output_file)
        if not output_path.exists():
            raise FileNotFoundError(f"Output file {self.output_file} does not exist")
        
        wb = load_workbook(output_path)
        diff_ws = wb[self.sheet_names['differences']]
        
        # Get labels (skip header row)
        labels = [cell.value for cell in diff_ws['A'][1:] if cell.value and cell.value != 'Label']
        
        # Create or update the average use sheet
        if self.sheet_names['average'] in wb.sheetnames:
            avg_ws = wb[self.sheet_names['average']]
            avg_ws.delete_rows(1, avg_ws.max_row)
        else:
            avg_ws = wb.create_sheet(self.sheet_names['average'])
        
        # Write header and labels
        avg_ws['A1'] = 'Label'
        for idx, label in enumerate(labels, start=2):
            avg_ws[f'A{idx}'] = label
        
        # Write header
        avg_ws['B1'] = 'Average Use'
        
        # Calculate averages
        for idx, label in enumerate(labels, start=2):
            differences = []
            for col in range(2, diff_ws.max_column + 1):
                col_letter = get_column_letter(col)
                val = diff_ws[f'{col_letter}{idx}'].value
                if val is not None:
                    try:
                        diff_value = float(val)
                        # Only include non-negative differences
                        if diff_value >= 0:
                            differences.append(diff_value)
                    except (ValueError, TypeError):
                        pass
            
            if differences:
                avg_use = sum(differences) / len(differences)
                avg_ws[f'B{idx}'] = round(avg_use, 2)
        
        # Organize sheets and remove defaults
        self._organize_sheets(wb)
        wb.save(output_path)
    
    def update_predictions(self, current_stock_file: str = None, current_stock_columns: tuple = ('Label', 'Stock')):
        """Update the Current Inventory & Predictions sheet.
        
        Creates predictions based on average use and compares with current stock.
        
        Args:
            current_stock_file: Path to file containing current stock data (if different from history)
            current_stock_columns: Tuple of (label_column, stock_column) header names
        """
        output_path = Path(self.output_file)
        if not output_path.exists():
            raise FileNotFoundError(f"Output file {self.output_file} does not exist")
        
        wb = load_workbook(output_path)
        avg_ws = wb[self.sheet_names['average']]
        history_ws = wb[self.sheet_names['history']]
        
        # Get current stock from the latest sale column in history
        labels = [cell.value for cell in avg_ws['A'][1:] if cell.value and cell.value != 'Label']
        latest_col = get_column_letter(history_ws.max_column)
        current_stock_map = {}
        
        for idx, label in enumerate(labels, start=2):
            stock = history_ws[f'{latest_col}{idx}'].value
            if stock is not None:
                try:
                    current_stock_map[label] = float(stock)
                except (ValueError, TypeError):
                    pass
        
        # Create or update the predictions sheet
        if self.sheet_names['predictions'] in wb.sheetnames:
            pred_ws = wb[self.sheet_names['predictions']]
            pred_ws.delete_rows(1, pred_ws.max_row)
        else:
            pred_ws = wb.create_sheet(self.sheet_names['predictions'])
        
        # Write headers
        pred_ws['A1'] = 'Label'
        pred_ws['B1'] = 'Current Stock'
        pred_ws['C1'] = 'Weekly Prediction'
        pred_ws['D1'] = 'Status'
        
        # Write data
        for idx, label in enumerate(labels, start=2):
            pred_ws[f'A{idx}'] = label
            
            # Current stock
            current = current_stock_map.get(label, 0)
            pred_ws[f'B{idx}'] = current
            
            # Average use * 7 for weekly prediction
            avg_use = avg_ws[f'B{idx}'].value
            if avg_use is not None:
                try:
                    prediction = float(avg_use) * 7
                    pred_ws[f'C{idx}'] = round(prediction, 2)
                    
                    # Status and coloring
                    if current >= prediction:
                        pred_ws[f'D{idx}'] = 'Adequate Stock'
                        pred_ws[f'D{idx}'].fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                        pred_ws[f'D{idx}'].font = Font(color='FFFFFF', bold=True)
                    else:
                        shortage = prediction - current
                        pred_ws[f'D{idx}'] = round(shortage, 2)
                        
                        # Color gradient from light green (1) to red (15+)
                        color = self._get_shortage_color(shortage)
                        pred_ws[f'D{idx}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                except (ValueError, TypeError):
                    pass
        
        # Adjust column widths
        pred_ws.column_dimensions['A'].width = 20
        pred_ws.column_dimensions['B'].width = 15
        pred_ws.column_dimensions['C'].width = 15
        pred_ws.column_dimensions['D'].width = 20
        
        # Organize sheets and remove defaults
        self._organize_sheets(wb)
        wb.save(output_path)
    
    def _get_shortage_color(self, shortage: float) -> str:
        """Generate a color based on shortage amount (light green to red gradient).
        
        Args:
            shortage: Amount of shortage
            
        Returns:
            Hex color code
        """
        # At 1: light green (C6EFCE), at 15+: red (FFC7CE)
        if shortage >= 15:
            return 'FFC7CE'
        
        # Linear interpolation between light green and red
        # Green: C6EFCE, Red: FFC7CE
        ratio = min(shortage / 15, 1.0)
        
        # RGB values
        r_start, g_start, b_start = 0xC6, 0xEF, 0xCE
        r_end, g_end, b_end = 0xFF, 0xC7, 0xCE
        
        r = int(r_start + (r_end - r_start) * ratio)
        g = int(g_start + (g_end - g_start) * ratio)
        b = int(b_start + (b_end - b_start) * ratio)
        
        return f'{r:02X}{g:02X}{b:02X}'
    
    def process_inventory(self, input_file: str, sale_number: str, 
                         label_column: str = 'Label', stock_column: str = 'Stock',
                         output_file: str = None) -> str:
        """Process inventory data end-to-end.
        
        Extracts data from input file and updates all analysis sheets.
        
        Args:
            input_file: Path to input Excel file
            sale_number: Sale number for this inventory
            label_column: Name of the column containing box labels
            stock_column: Name of the column containing stock values
            output_file: Optional custom output file path
            
        Returns:
            Path to the generated analysis file
            
        Raises:
            FileNotFoundError: If input file doesn't exist
            ValueError: If required columns aren't found
        """
        if output_file:
            self.output_file = output_file
        
        input_path = Path(input_file)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")
        
        # Extract data
        data = self.extract_columns(input_file, [label_column, stock_column])
        labels = data[label_column]
        stock = data[stock_column]
        
        # Update all sheets
        self.update_inventory_history(labels, stock, sale_number)
        self.update_sales_differences()
        self.update_average_use()
        self.update_predictions()
        
        return str(Path(self.output_file).resolve())
    
    def process_restock(self, input_file: str, 
                       label_column: str = 'Label', stock_column: str = 'Stock',
                       output_file: str = None) -> str:
        """Process inventory restock data end-to-end.
        
        Extracts data from input file and adds it as a restock entry without calculating
        differences from this restock to the previous entry (but differences to the next sale
        will be calculated if they are consecutive).
        
        Args:
            input_file: Path to input Excel file
            label_column: Name of the column containing box labels
            stock_column: Name of the column containing stock values
            output_file: Optional custom output file path
            
        Returns:
            Path to the generated analysis file
            
        Raises:
            FileNotFoundError: If input file doesn't exist
            ValueError: If required columns aren't found
        """
        if output_file:
            self.output_file = output_file
        
        input_path = Path(input_file)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")
        
        # Extract data
        data = self.extract_columns(input_file, [label_column, stock_column])
        labels = data[label_column]
        stock = data[stock_column]
        
        # Update inventory history with restock entry
        self._update_inventory_history_restock(labels, stock)
        self.update_sales_differences()
        self.update_average_use()
        self.update_predictions()
        
        return str(Path(self.output_file).resolve())
